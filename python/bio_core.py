#!/usr/bin/env python3
"""
BioViz Local - Python Sidecar (Long-Running Daemon Process)

This is NOT a one-shot script. It runs as a persistent daemon process,
communicating with the Rust backend via STDIO (stdin/stdout).

Protocol:
- Read one JSON line from stdin
- Process the command
- Write one JSON line to stdout (immediately flushed)
- Repeat forever until killed

Commands:
- {"cmd": "HEARTBEAT"} -> {"status": "alive"}
- {"cmd": "load", "path": "..."} -> {"status": "ok", "data": ...}
- {"cmd": "analyze", ...} -> {"status": "ok", "result": ...}
"""

import argparse
import sys
import os
import json
import logging
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass
import traceback
import math
from pathlib import Path
from typing import Dict, Any, Optional, List, Tuple
from mapper import color_kegg_pathway, get_pathway_statistics

# Import v2.0 modules
try:
    from gsea_module import (
        handle_run_enrichr,
        handle_run_gsea,
        handle_get_gene_sets,
        check_gsea_available
    )
    GSEA_AVAILABLE = check_gsea_available()
    logging.info(f"[INIT] GSEA module imported, available={GSEA_AVAILABLE}")
except ImportError as e:
    GSEA_AVAILABLE = False
    logging.warning(f"[INIT] GSEA module not available: {e}")
    print("[BioEngine] GSEA module not available", file=sys.stderr)

try:
    from image_module import (
        handle_upload_image,
        handle_analyze_image,
        handle_list_images,
        check_image_available
    )
    IMAGE_AVAILABLE = check_image_available()
    logging.info(f"[INIT] Image module imported, available={IMAGE_AVAILABLE}")
except ImportError as e:
    IMAGE_AVAILABLE = False
    logging.warning(f"[INIT] Image module not available: {e}")
    print("[BioEngine] Image module not available", file=sys.stderr)

try:
    from multi_sample import (
        handle_load_multi_sample,
        handle_get_sample_groups
    )
    MULTI_SAMPLE_AVAILABLE = True
    logging.info("[INIT] Multi-sample module imported successfully")
except ImportError as e:
    MULTI_SAMPLE_AVAILABLE = False
    logging.warning(f"[INIT] Multi-sample module not available: {e}")
    print("[BioEngine] Multi-sample module not available", file=sys.stderr)

# Import pathway template manager
try:
    from pathway.template_manager import PathwayTemplateManager
    TEMPLATE_MANAGER = PathwayTemplateManager()
    logging.info("[INIT] PathwayTemplateManager initialized")
except ImportError as e:
    TEMPLATE_MANAGER = None
    logging.warning(f"[INIT] Template manager not available: {e}")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Ensure stdout is line-buffered for immediate response
# This is CRITICAL for real-time IPC
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(line_buffering=True)

# --- Request context (for correlating async frontend calls) ---
# The engine is single-threaded, so a simple global context works.
CURRENT_REQUEST_ID: Optional[str] = None
CURRENT_CMD: Optional[str] = None


def send_response(data: Dict[str, Any]) -> None:
    """Send a JSON response to stdout and flush immediately."""
    try:
        # Attach request correlation info when available.
        # This keeps backward compatibility for clients that ignore these fields.
        if CURRENT_REQUEST_ID is not None and "request_id" not in data:
            data["request_id"] = CURRENT_REQUEST_ID
        if CURRENT_CMD is not None and "cmd" not in data:
            data["cmd"] = CURRENT_CMD
        json_str = json.dumps(data, ensure_ascii=False)
        print(json_str, flush=True)
    except Exception as e:
        # Fallback error response
        error_response = json.dumps({"status": "error", "message": str(e)})
        print(error_response, flush=True)


def send_error(message: str, details: Dict[str, Any] = None) -> None:
    """Send an error response."""
    data = {"status": "error", "message": message}
    if details:
        data.update(details)
    send_response(data)


def handle_heartbeat(payload: Dict[str, Any] = None) -> Dict[str, Any]:
    """Handle HEARTBEAT command - used for health checks."""
    return {"status": "alive"}


def preprocess_matrix_if_needed(file_path: str) -> str:
    """
    Detects if the file is a 'Wide' Omics Matrix (Genes as columns) and converts it 
    to a standard 'Long' format (Gene, Value) for the tool.
    
    Returns:
        str: Path to the processed file (temp file) or original path if no processing needed.
    """
    try:
        if not file_path.lower().endswith(('.csv', '.txt', '.tsv')):
            return file_path

        import csv
        import re
        import tempfile
        
        delimiter = '\t' if file_path.lower().endswith('.tsv') else ','
        
        # Heuristics for "Wide Matrix"
        # 1. Many columns (> 50)
        # 2. Header contains Gene-like names (e.g., "A1BG", "TP53 (1234)")
        
        with open(file_path, 'r', encoding='utf-8-sig', errors='replace') as f:
            # Check first line
            line = f.readline()
            if not line: return file_path
            
            # Reset seek to sniff properly or just parse line
            # We'll re-open or seek 0 later.
            
            # Simple parse of first line
            headers = [h.strip() for h in line.split(delimiter)]
            col_count = len(headers)
            
            if col_count < 50:
                return file_path
                
            # Check for Gene pattern "Symbol (ID)" or just "Symbol"
            # If >50 cols, highly likely omics.
            # Let's perform the transposition.
            print(f"[BioEngine] High column count ({col_count}) detected. Attempting transpose...", file=sys.stderr)
            
            # Read the first data row (Sample 1)
            line2 = f.readline()
            if not line2:
                return file_path # Only headers?
                
            values = [v.strip() for v in line2.split(delimiter)]
            
            # Ensure length match
            if len(values) != len(headers):
               # Handle ragged CSVs gracefully?
               # For now, simplistic approach: truncate to shorted
               min_len = min(len(values), len(headers))
               headers = headers[:min_len]
               values = values[:min_len]

        # Prepare Transposed Data
        # We assume Col 0 might be "index" or empty, check if Header[0] is Gene
        # Usually in R output: "", "Gene1", "Gene2"...
        # Row1: "Sample1", 0.1, 0.2...
        
        # If Header[0] is empty, skip it.
        # If Value[0] is string (Sample Name), skip it.
        
        start_idx = 0
        try:
            float(values[start_idx])
        except ValueError:
            # First col is likely sample name
            start_idx = 1
            
        # Clean Gene Names
        # Regex to remove (EntrezID) e.g., "TP53 (7157)" -> "TP53"
        clean_rows = []
        for i in range(start_idx, len(headers)):
            gene_raw = headers[i]
            # Match "Gene (123)"
            m = re.match(r"(.+?)\s*\(\d+\)", gene_raw)
            gene_clean = m.group(1) if m else gene_raw
            gene_clean = gene_clean.replace('"', '').strip() # Remove quotes if any
            
            # Skip empty genes
            if not gene_clean: continue
            
            try:
                val = float(values[i])
                clean_rows.append([gene_clean, val])
            except ValueError:
                continue # Skip non-numeric values
                
        if not clean_rows:
            return file_path # Failed to extract data
            
        # Write to Temp File
        temp_fd, temp_path = tempfile.mkstemp(suffix='_transposed.csv', prefix='bioviz_')
        os.close(temp_fd)
        
        with open(temp_path, 'w', encoding='utf-8', newline='') as tf:
            writer = csv.writer(tf)
            writer.writerow(['Gene', 'Value']) # Standard Header
            writer.writerows(clean_rows)
            
        print(f"[BioEngine] Transposed matrix saved to: {temp_path}", file=sys.stderr)
        return temp_path

    except Exception as e:
        print(f"[BioEngine] Pre-process failed: {e}", file=sys.stderr)
        return file_path


# --- Helper functions for analysis ---

LOGFC_KEYWORDS = [
    'logfc', 'log2fc', 'log fc', 'log2 fold', 'log2fold', 'fold change',
    'foldchange', 'fc', 'log2ratio', 'log2_ratio', 'abundanceratio',
]

CONTROL_PATTERNS = [
    'control', 'ctrl', 'cond_a', 'condition_a', 'condition a', 'baseline', 'untreated'
]

TREAT_PATTERNS = [
    'treat', 'treated', 'cond_b', 'condition_b', 'condition b',
    'stim', 'case', 'exp', 'expr', 'experiment', 'experimental', 'trt'
]

def _guess_gene_header(headers: List[str]) -> Optional[str]:
    """
    Try to guess the gene/protein/cell column by header keywords.
    This is a lightweight fallback when the user-selected column
    is missing in a specific file (e.g., GeneName vs Gene).
    """
    if not headers:
        return None
    candidates: List[str] = []
    keywords = ['gene', 'symbol', 'name', 'id', 'identifier', 'accession', 'cell', 'protein', 'uniprot']
    for h in headers:
        h_lower = str(h).lower()
        if any(k in h_lower for k in keywords):
            candidates.append(h)
    if candidates:
        # Prefer columns that explicitly mention gene/symbol/name
        def _score(col: str) -> int:
            l = col.lower()
            score = 0
            if 'gene' in l or 'symbol' in l:
                score += 3
            if 'name' in l:
                score += 2
            if 'id' in l or 'accession' in l:
                score -= 1
            return score
        candidates.sort(key=_score, reverse=True)
        return candidates[0]
    return None


def _guess_value_header(headers: List[str]) -> Optional[str]:
    """
    Guess the value / logFC column by header keywords.
    """
    if not headers:
        return None
    keywords = [
        'logfc', 'log2fc', 'log fc', 'log2 fold', 'fold change', 'foldchange', 'fc',
        'ratio', 'log2ratio', 'log2_ratio', 'expression', 'expr', 'value', 'intensity',
        'score', 'abundance'
    ]
    for h in headers:
        h_lower = str(h).lower()
        if any(k in h_lower for k in keywords):
            return h
    return None


def looks_like_logfc(col_name: Optional[str]) -> bool:
    """Heuristic: column name suggests it already stores log2 fold-change."""
    if not col_name:
        return False
    lower = col_name.lower()
    return any(key in lower for key in LOGFC_KEYWORDS)


def infer_control_treat_indices(headers: List[str], gene_idx: int) -> Tuple[List[int], List[int]]:
    """
    Infer replicate groups from header names.
    Very lightweight: any column whose name contains 'control' goes to control group,
    'treat' / 'treated' goes to treatment group, plus a few Condition_A/B aliases.
    """
    control_idx: List[int] = []
    treat_idx: List[int] = []
    for i, h in enumerate(headers):
        if i == gene_idx:
            continue
        lower = (h or '').lower()
        if any(p in lower for p in CONTROL_PATTERNS):
            control_idx.append(i)
        if any(p in lower for p in TREAT_PATTERNS):
            treat_idx.append(i)
    return control_idx, treat_idx


def normal_cdf(z: float) -> float:
    """Standard normal CDF using error function."""
    return 0.5 * (1.0 + math.erf(z / math.sqrt(2.0)))


def variance(values: List[float]) -> float:
    """Sample variance (unbiased, n-1 in denominator)."""
    n = len(values)
    if n < 2:
        return 0.0
    mean_val = sum(values) / n
    return sum((v - mean_val) ** 2 for v in values) / (n - 1)


def write_analysis_table(original_path: str, volcano_data: List[Dict[str, Any]]) -> Optional[str]:
    """
    将火山图数据写出为一个独立的统计结果表，用于“留证据”。
    输出格式统一为 CSV，放在原始文件同一目录下，文件名加后缀 *_bioviz_stats.csv。
    """
    if not volcano_data:
        return None
    try:
        abs_path = os.path.abspath(original_path)
        base, _ext = os.path.splitext(abs_path)
        # Avoid stacking suffixes if input is already a *_bioviz_stats.csv file
        if base.endswith("_bioviz_stats"):
            out_path = abs_path
        else:
            out_path = f"{base}_bioviz_stats.csv"
        import csv
        with open(out_path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Entity', 'Log2FC', '-log10(P)', 'PValue', 'Status'])
            for row in volcano_data:
                writer.writerow([
                    row.get('gene', ''),
                    row.get('x', ''),
                    row.get('y', ''),
                    row.get('pvalue', ''),
                    row.get('status', ''),
                ])
        print(f"[BioEngine] Analysis table written to: {out_path}", file=sys.stderr)
        return out_path
    except Exception as e:
        print(f"[BioEngine] Failed to write analysis table: {e}", file=sys.stderr)
        return None


def handle_load(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Handle load command - load Excel/CSV file using standard libraries."""
    original_path = payload.get("path", "")
    if not original_path:
        return {"status": "error", "message": "Missing 'path' parameter"}
    
    # 1. Pre-process (Transpose if Wide Matrix)
    path = preprocess_matrix_if_needed(original_path)
    is_processed = path != original_path
    
    try:
        columns: List[str] = []
        preview: List[List[str]] = []
        total_rows: Optional[int] = None
        
        # Read file based on extension
        if path.lower().endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb.active

                rows_iter = ws.iter_rows(values_only=True)
                header_row = next(rows_iter, None)
                if not header_row:
                    return {"status": "error", "message": "File is empty"}

                columns = [str(c) if c is not None else '' for c in header_row]
                for _ in range(5):
                    row = next(rows_iter, None)
                    if row is None:
                        break
                    preview.append([str(cell) if cell is not None else '' for cell in row])

                # Avoid scanning the entire workbook for row count (can be slow for large files).
                try:
                    if isinstance(ws.max_row, int) and ws.max_row >= 1:
                        total_rows = max(ws.max_row - 1, 0)
                except Exception:
                    total_rows = None
                
            except ImportError:
                 return {"status": "error", "message": "openpyxl module not found. Please pip install openpyxl"}
                 
        elif path.lower().endswith(('.csv', '.txt', '.tsv')):
            import csv
            
            # Detect delimiter (simple check)
            delimiter = ','
            if path.lower().endswith('.tsv'):
                delimiter = '\t'
            
            with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
                # Read sample to sniff delimiter if needed, but simple approach first
                # Check line 1 for delimiter
                first_line = f.readline()
                f.seek(0)
                if ';' in first_line and ',' not in first_line:
                    delimiter = ';'
                
                reader = csv.reader(f, delimiter=delimiter)
                try:
                    columns = next(reader)
                except StopIteration:
                     return {"status": "error", "message": "File is empty"}
                     
                def _clean_header(val: Any) -> str:
                    try:
                        s = str(val).strip()
                        # Remove wrapping quotes/backticks
                        s = s.strip('"').strip("'").strip('`')
                        return s
                    except Exception:
                        return str(val)

                columns = [_clean_header(c) for c in columns]
                 
                # Read first 5 rows for preview
                preview = []
                for _ in range(5):
                    try:
                        row = next(reader)
                        preview.append(row)
                    except StopIteration:
                        break

                # Avoid scanning the entire file for row count (can be very slow for large CSV/TSV).
                total_rows = None

        else:
            return {"status": "error", "message": "Unsupported file format. Use .xlsx, .xls, .csv, or .tsv"}
        
        # --- Smart column inference (regex + rules) ---
        gene_column = None
        value_column = None
        pvalue_column = None
        group_column = None

        # Extended keywords for Entity column (Gene, Protein, Cell)
        entity_keywords = [
            # Generic/Gene
            'gene', 'symbol', 'id', 'identifier', 'accession', 'entity',
            'cellpopulation', 'population',
            # Cell
            'cell type', 'cellname', 'cell_name', 'cell', 'lineage', 'phenotype', 'cluster',
            # Protein
            'protein', 'uniprot', 'peptide'
        ]

        # Keywords for LogFC/Value column (exclude P-value related)
        value_keywords_priority = [
            # High priority: Expression/LogFC (most common for differential expression)
            'logfc', 'log2fc', 'log fc', 'log2 fold', 'fold change', 'foldchange', 'fc', 'log2', 'log_fc',
            'expression', 'expr', 'ratio', 'log2ratio', 'abundanceratio',
            # Medium priority: General values
            'value', 'score', 'intensity',
            # Lower priority: Cell/Flow cytometry metrics
            'frequency', 'freq', 'count', 'abundance', 'percent', 'percentage', 'proportion'
        ]

        # Keywords for P-value column (separate from value)
        pvalue_keywords = [
            'pvalue', 'p-value', 'p.value', 'pval', 'p_value',
            'adj.p.val', 'padj', 'fdr', 'q-value', 'qvalue',
            'significance', 'adj_pvalue'
        ]

        # Keywords for group/condition columns (for potential raw-data support)
        group_keywords = [
            'group', 'condition', 'treatment', 'timepoint', 'time', 'batch', 'sampletype'
        ]

        # 1. Find Entity Column (score-based, prefer Symbol/Name over ID)
        entity_candidates: List[str] = []
        for col in columns:
            col_lower = col.lower()
            if any(k in col_lower for k in entity_keywords):
                entity_candidates.append(col)

        if entity_candidates:
            def _entity_score(c: str) -> int:
                l = c.lower()
                score = 0
                if 'symbol' in l or 'name' in l or 'celltype' in l:
                    score += 3
                if 'gene' in l or 'protein' in l or 'cell' in l:
                    score += 2
                if 'id' in l or 'accession' in l:
                    score -= 1
                return score

            entity_candidates.sort(key=_entity_score, reverse=True)
            gene_column = entity_candidates[0]

        # 2. Find Value Column (prioritize LogFC/ratio, exclude P-value columns)
        for col in columns:
            if col == gene_column:
                continue
            col_lower = col.lower()
            # Skip if this looks like a P-value column
            if any(k in col_lower for k in pvalue_keywords):
                continue
            if any(k in col_lower for k in value_keywords_priority):
                value_column = col
                break

        # 3. Find P-value Column (for volcano plot support)
        for col in columns:
            if col == gene_column or col == value_column:
                continue
            col_lower = col.lower()
            if any(k in col_lower for k in pvalue_keywords):
                pvalue_column = col
                break

        # 4. Try to find a group/condition column (non-numeric, few unique values)
        if preview:
            # Use first few rows for heuristics
            max_rows = min(len(preview), 20)
            for idx, col in enumerate(columns):
                # Skip obvious numeric candidates (value / pvalue)
                if col in (gene_column, value_column, pvalue_column):
                    continue
                col_lower = col.lower()
                # Prefer keyword matches
                keyword_hit = any(k in col_lower for k in group_keywords)

                values = []
                numeric_count = 0
                for r in range(max_rows):
                    row = preview[r]
                    if idx >= len(row):
                        continue
                    v = str(row[idx]).strip()
                    if v == '':
                        continue
                    values.append(v)
                    try:
                        float(v)
                        numeric_count += 1
                    except ValueError:
                        pass

                if not values:
                    continue

                # If majority numeric, likely not a group column
                if numeric_count >= len(values) * 0.6:
                    continue

                distinct_vals = len(set(values))
                # Heuristic: group-like if categories are few (2–10)
                if 1 < distinct_vals <= 10 or keyword_hit:
                    group_column = col
                    break

        # Build suggested mapping for the front-end wizard
        suggested_mapping: Dict[str, Any] = {}
        if gene_column:
            suggested_mapping['gene'] = gene_column
        if value_column:
            suggested_mapping['value'] = value_column
        if pvalue_column:
            suggested_mapping['pvalue'] = pvalue_column

        # --- Layout / data-type inference (for diagnostics & future features) ---
        # Classify table shape: summary vs long_raw vs wide_matrix
        layout = "unknown"
        if gene_column and value_column and pvalue_column:
            layout = "summary"
        elif gene_column and value_column and group_column:
            layout = "long_raw"
        elif len(columns) >= 50 or is_processed:
            layout = "wide_matrix"
        else:
            layout = "summary_like"

        # Very lightweight guess of biological data type from header text
        joined_headers = " ".join(c.lower() for c in columns)
        data_type_guess = "gene"
        if any(k in joined_headers for k in ['protein', 'uniprot', 'phospho']):
            data_type_guess = "protein"
        elif any(k in joined_headers for k in ['cell type', 'flow', 'fcs']):
            data_type_guess = "cell"

        message = "Successfully loaded file"
        if isinstance(total_rows, int):
            message = f"Successfully loaded {total_rows} rows"

        return {
            "status": "ok",
            "message": message,
            "path": path,  # Return the potentially modified path
            "rows": total_rows,
            "columns": columns,
            "preview": preview,
            "suggested_mapping": suggested_mapping,
            "is_transposed": is_processed,
            "data_layout": layout,
            "data_type_guess": data_type_guess,
            "group_column_guess": group_column,
        }
    except Exception as e:
        import traceback
        return {"status": "error", "message": f"Failed to load file: {str(e)}", "traceback": traceback.format_exc()}


def handle_analyze(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Handle analyze command - color KEGG pathway with gene expression data and generate volcano plot data.

    支持两种数据来源：
    1. Summary：映射列已经是 log2FC / P-value。
    2. Raw matrix：存在 Control_*/Treat_* 等重复列，自动计算 log2FC + 近似 P-value。
    """
    file_path = payload.get("file_path", "")
    mapping = payload.get("mapping", {})
    template_id = payload.get("template_id", "")
    data_type = payload.get("data_type", "gene")
    filters = payload.get("filters", {})  # Optional: {pvalue_threshold, logfc_threshold, method, methods}
    
    if not file_path:
        return {"status": "error", "message": "Missing 'file_path' parameter"}
    if not mapping or 'gene' not in mapping or 'value' not in mapping:
        return {"status": "error", "message": "Missing or invalid 'mapping' parameter"}
    # template_id is now optional in the 3-step workflow
    
    try:
        gene_col = mapping['gene']
        value_col = mapping['value']
        pvalue_col = mapping.get('pvalue')  # Optional P-value column
        control_cols = mapping.get("controlCols") or mapping.get("control_cols") or []
        treat_cols = mapping.get("treatCols") or mapping.get("treat_cols") or []
        # Normalize list of control/experiment columns if provided
        control_cols = [str(c) for c in control_cols if isinstance(c, (str, int))]
        treat_cols = [str(c) for c in treat_cols if isinstance(c, (str, int))]
        
        # Filter thresholds with defaults
        pvalue_threshold = float(filters.get('pvalue_threshold', 0.05))
        logfc_threshold = float(filters.get('logfc_threshold', 1.0))
        # Primary analysis method (single) used to drive downstream logic.
        # Frontend may also send filters["methods"] as a list for multi-select UX,
        # but the engine currently uses only this primary method.
        methods_list = filters.get('methods')
        if isinstance(methods_list, (list, tuple)) and methods_list:
            method = str(methods_list[0]).lower()
        else:
            method = str(filters.get('method', 'auto')).lower()
        
        gene_expression: Dict[str, float] = {}
        gene_pvalues: Dict[str, float] = {}  # Store P-values for volcano plot
        # Optional per-entity mean expression for MA plot (A-value)
        # When available, this will drive the X 轴 of the MA 图。
        gene_means: Dict[str, float] = {}
        
        # Read file and extract data
        if file_path.lower().endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                
                if not rows:
                     return {"status": "error", "message": "File is empty"}
                
                headers = [str(c) if c is not None else '' for c in rows[0]]
                try:
                    gene_idx = headers.index(gene_col)
                except ValueError:
                    guessed = _guess_gene_header(headers)
                    if guessed and guessed in headers:
                        gene_idx = headers.index(guessed)
                        print(f"[BioEngine] Gene column '{gene_col}' not found. Using '{guessed}' instead.", file=sys.stderr)
                    else:
                        return {"status": "error", "message": f"Gene column '{gene_col}' not found in headers: {headers}"}

                try:
                    value_idx = headers.index(value_col)
                except ValueError:
                    guessed_val = _guess_value_header(headers)
                    if guessed_val and guessed_val in headers:
                        value_idx = headers.index(guessed_val)
                        print(f"[BioEngine] Value column '{value_col}' not found. Using '{guessed_val}' instead.", file=sys.stderr)
                    else:
                        return {"status": "error", "message": f"Value column '{value_col}' not found in headers: {headers}"}

                pvalue_idx = headers.index(pvalue_col) if pvalue_col and pvalue_col in headers else None

                # XLS(X) 目前仅支持 summary 模式
                for row in rows[1:]:
                    if len(row) <= max(gene_idx, value_idx):
                        continue

                    gene = str(row[gene_idx]).strip()
                    try:
                        val_raw = row[value_idx]
                        if val_raw is not None:
                            val = float(val_raw)
                            if gene:
                                gene_expression[gene] = val
                                if pvalue_idx is not None and len(row) > pvalue_idx:
                                    try:
                                        pval = float(row[pvalue_idx] or 0.0)
                                        gene_pvalues[gene] = pval
                                    except (ValueError, TypeError):
                                        pass
                    except (ValueError, TypeError):
                        continue
                        
            except ImportError:
                return {"status": "error", "message": "openpyxl module not found"}
                
        elif file_path.lower().endswith(('.csv', '.txt', '.tsv')):
            import csv
            delimiter = '\t' if file_path.lower().endswith('.tsv') else ','
            
            with open(file_path, 'r', encoding='utf-8-sig', errors='replace') as f:
                reader = csv.reader(f, delimiter=delimiter)
                try:
                    headers = next(reader)
                except StopIteration:
                    return {"status": "error", "message": "File is empty"}

                def _clean_header(val: Any) -> str:
                    try:
                        s = str(val).strip()
                        s = s.strip('"').strip("'").strip('`')
                        return s
                    except Exception:
                        return str(val)

                headers = [_clean_header(h) for h in headers]

                try:
                    gene_idx = headers.index(gene_col)
                except ValueError:
                    guessed = _guess_gene_header(headers)
                    if guessed and guessed in headers:
                        gene_idx = headers.index(guessed)
                        print(f"[BioEngine] Gene column '{gene_col}' not found. Using '{guessed}' instead.", file=sys.stderr)
                    else:
                        return {"status": "error", "message": f"Gene column '{gene_col}' not found in headers: {headers}"}

                # Optional indices for summary-mode mapping
                if value_col in headers:
                    value_idx = headers.index(value_col)
                else:
                    guessed_val = _guess_value_header(headers)
                    if guessed_val and guessed_val in headers:
                        value_idx = headers.index(guessed_val)
                        print(f"[BioEngine] Value column '{value_col}' not found. Using '{guessed_val}' instead.", file=sys.stderr)
                    else:
                        value_idx = None

                pvalue_idx = headers.index(pvalue_col) if pvalue_col and pvalue_col in headers else None

                # Optional mean-expression column for summary tables (e.g. DESeq2 BaseMean)
                mean_idx: Optional[int] = None
                lowered_headers = [h.lower() for h in headers]
                for i, h in enumerate(lowered_headers):
                    # 专门识别 DESeq2 风格的 BaseMean / baseMean
                    if h.replace(" ", "").replace("_", "") in ("basemean", "base.mean"):
                        mean_idx = i
                        break

                # Detect replicate groups for potential Raw matrix mode
                if control_cols and treat_cols:
                    control_idx = [headers.index(c) for c in control_cols if c in headers]
                    treat_idx = [headers.index(c) for c in treat_cols if c in headers]
                    if not control_idx or not treat_idx:
                        return {
                            "status": "error",
                            "message": f"Selected control/experiment columns not found in headers: {headers}"
                        }
                else:
                    control_idx, treat_idx = infer_control_treat_indices(headers, gene_idx)

                if method == "precomputed":
                    use_raw_mode = False
                elif method == "ttest":
                    use_raw_mode = bool(control_idx and treat_idx)
                else:  # auto
                    use_raw_mode = (
                        control_idx
                        and treat_idx
                        and not pvalue_idx
                        and (
                            value_col == '__raw_matrix__'
                            or not looks_like_logfc(value_col)
                        )
                    )

                if use_raw_mode:
                    # --- Raw matrix mode: Control_* / Treat_* 列，多重复 ---
                    eps = 1e-6
                    for row in reader:
                        if len(row) <= gene_idx:
                            continue

                        gene = row[gene_idx].strip()
                        if not gene:
                            continue

                        control_vals: List[float] = []
                        treat_vals: List[float] = []

                        for idx in control_idx:
                            if idx < len(row):
                                v = row[idx].strip()
                                if v != '':
                                    try:
                                        control_vals.append(float(v))
                                    except (ValueError, TypeError):
                                        continue

                        for idx in treat_idx:
                            if idx < len(row):
                                v = row[idx].strip()
                                if v != '':
                                    try:
                                        treat_vals.append(float(v))
                                    except (ValueError, TypeError):
                                        continue

                        if not control_vals or not treat_vals:
                            continue

                        mean_c = sum(control_vals) / len(control_vals)
                        mean_t = sum(treat_vals) / len(treat_vals)

                        ratio = (mean_t + eps) / (mean_c + eps)
                        logfc = math.log2(ratio)

                        # MA plot uses平均表达量 (A 值) 作为 X 轴：
                        # A = 0.5 * (log2(mean_t) + log2(mean_c))
                        ma_mean = 0.5 * (
                            math.log2(mean_t + eps) + math.log2(mean_c + eps)
                        )

                        # 近似 P 值：正态近似的双尾检验
                        if len(control_vals) > 1 and len(treat_vals) > 1:
                            var_c = variance(control_vals)
                            var_t = variance(treat_vals)
                            se = math.sqrt(var_c / len(control_vals) + var_t / len(treat_vals))
                            if se > 0:
                                z = (mean_t - mean_c) / se
                                pval = 2.0 * (1.0 - normal_cdf(abs(z)))
                            else:
                                pval = 1.0
                        else:
                            pval = 1.0

                        gene_expression[gene] = logfc
                        gene_pvalues[gene] = pval
                        gene_means[gene] = ma_mean
                else:
                    # --- Summary 模式：mapping.value 已经是 logFC 或评分 ---
                    if value_idx is None:
                        return {
                            "status": "error",
                            "message": f"Value column '{value_col}' not found in headers: {headers}",
                        }

                    for row in reader:
                        if len(row) <= max(gene_idx, value_idx):
                            continue

                        gene = row[gene_idx].strip()
                        try:
                            val = float(row[value_idx])
                            if gene:
                                gene_expression[gene] = val
                                if pvalue_idx is not None and len(row) > pvalue_idx:
                                    try:
                                        pval = float(row[pvalue_idx])
                                        gene_pvalues[gene] = pval
                                    except (ValueError, TypeError):
                                        pass
                                # Optional mean column for MA plot
                                if mean_idx is not None and len(row) > mean_idx:
                                    try:
                                        mean_val = float(row[mean_idx])
                                        gene_means[gene] = mean_val
                                    except (ValueError, TypeError):
                                        pass
                        except (ValueError, TypeError):
                            continue

        if not gene_expression:
            return {"status": "error", "message": "No valid gene expression data found"}
        
        # Generate Volcano Plot Data
        volcano_data: List[Dict[str, Any]] = []
        for gene, logfc in gene_expression.items():
            pval = gene_pvalues.get(gene, 1.0)  # Default to 1.0 (not significant) if no P-value
            mean_val = gene_means.get(gene)
            
            # Calculate -log10(pvalue) for Y-axis, handle edge cases
            if pval <= 0:
                neg_log_pval = 10.0  # Cap at 10 for p-value of 0
            elif pval >= 1:
                neg_log_pval = 0.0
            else:
                neg_log_pval = -math.log10(pval)
            
            # Determine status: UP, DOWN, or NS (not significant)
            is_significant = pval < pvalue_threshold
            is_up = logfc > logfc_threshold
            is_down = logfc < -logfc_threshold
            
            if is_significant and is_up:
                status = "UP"
            elif is_significant and is_down:
                status = "DOWN"
            else:
                status = "NS"
            
            row: Dict[str, Any] = {
                "gene": gene,
                "x": round(logfc, 4),
                "y": round(neg_log_pval, 4),
                "pvalue": pval,
                "status": status
            }
            if mean_val is not None:
                row["mean"] = round(mean_val, 4)
            volcano_data.append(row)

        # Persist analysis table next to original data for traceability
        # analysis_table_path = write_analysis_table(file_path, volcano_data)
        analysis_table_path = None

        # Color the pathway if template_id is provided
        if template_id:
            colored_pathway = color_kegg_pathway(template_id, gene_expression, data_type=data_type)
            statistics = get_pathway_statistics(colored_pathway)
        else:
            colored_pathway = None
            # Generic statistics fallback when no pathway is selected
            up = len([v for v in volcano_data if v['status'] == 'UP'])
            down = len([v for v in volcano_data if v['status'] == 'DOWN'])
            ns = len(volcano_data) - up - down
            statistics = {
                'total_nodes': len(volcano_data),
                'upregulated': up,
                'downregulated': down,
                'unchanged': ns,
                'percent_upregulated': 100 * up / len(volcano_data) if volcano_data else 0,
                'percent_downregulated': 100 * down / len(volcano_data) if volcano_data else 0
            }
        
        # Generate AI insights from analysis results
        try:
            from tools.insight_generator import generate_insights
            analysis_result = {
                "statistics": statistics,
                "volcano_data": volcano_data,
                "gene_count": len(gene_expression),
                "has_pvalue": len(gene_pvalues) > 0
            }
            insights = generate_insights(analysis_result)
        except Exception as e:
            print(f"[BioEngine] Failed to generate insights: {e}", file=sys.stderr)
            insights = {"summary": "", "badges": []}
        
        # Return both pathway and volcano data with AI insights
        return {
            "status": "ok",
            "pathway": colored_pathway,
            "statistics": statistics,
            "gene_count": len(gene_expression),
            "volcano_data": volcano_data,
            "has_pvalue": len(gene_pvalues) > 0,
            "analysis_table_path": analysis_table_path,
            "insights": insights,  # AI-generated insights
        }
    except Exception as e:
        import traceback
        return {
            "status": "error",
            "message": f"Analysis failed: {str(e)}",
            "traceback": traceback.format_exc()
        }


def handle_load_pathway(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Handle load_pathway command - load a KEGG pathway template."""
    pathway_id = payload.get("pathway_id", "")
    if not pathway_id:
        return {"status": "error", "message": "Missing 'pathway_id' parameter"}
    
    try:
        from mapper import load_pathway_template
        pathway = load_pathway_template(pathway_id)
        
        if not pathway:
            return {
                "status": "error",
                "message": f"Pathway template '{pathway_id}' not found"
            }
        
        return {
            "status": "ok",
            "pathway": pathway
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


def handle_color_pathway(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Handle color_pathway command - apply gene expression coloring."""
    pathway_id = payload.get("pathway_id", "")
    gene_expression = payload.get("gene_expression", {})
    
    if not pathway_id:
        return {"status": "error", "message": "Missing 'pathway_id' parameter"}
    
    if not gene_expression:
        return {"status": "error", "message": "Missing 'gene_expression' data"}
    
    try:
        colored_pathway = color_kegg_pathway(pathway_id, gene_expression)
        statistics = get_pathway_statistics(colored_pathway)
        
        return {
            "status": "ok",
            "pathway": colored_pathway,
            "statistics": statistics
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


def handle_save_data(payload: Dict[str, Any]) -> None:
    """Save raw analysis data (volcano data) to a CSV file."""
    try:
        path = payload.get("path", "")
        data = payload.get("data", [])
        
        if not path:
             send_error("Missing 'path' parameter")
             return
             
        if not data:
             send_error("Missing 'data' parameter")
             return

        import csv
        with open(path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            # Write header
            # NOTE: Use a header that does not start with '-' or '='
            # so that Excel does not auto-interpret it as a formula.
            writer.writerow(['Gene', 'Log2FC', 'neg_log10(P)', 'PValue', 'Status', 'Mean'])
            
            for row in data:
                writer.writerow([
                    row.get('gene', ''),
                    row.get('x', ''),
                    row.get('y', ''),
                    row.get('pvalue', ''),
                    row.get('status', ''),
                    row.get('mean', '')
                ])
                
        send_response({
            "status": "ok", 
            "message": "Data saved successfully",
            "path": path
        })
    except Exception as e:
        send_error(f"Failed to save data: {str(e)}")


def handle_save_analysis(payload: Dict[str, Any]) -> None:
    """Save current analysis to a local file"""
    try:
        from pathlib import Path
        import time
        
        # Get user home directory for persistence
        home_dir = Path.home() / '.bioviz_local'
        home_dir.mkdir(exist_ok=True)
        
        timestamp = int(time.time())
        filename = f"analysis_{timestamp}.json"
        filepath = home_dir / filename
        
        data = {
            "timestamp": timestamp,
            "pathway_id": payload.get("pathway_id"),
            "gene_expression": payload.get("gene_expression"),
            "pathway_data": payload.get("pathway_data"),
            "statistics": payload.get("statistics")
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f)
            
        send_response({
            "status": "ok", 
            "message": "Analysis saved successfully",
            "filename": filename,
            "filepath": str(filepath)
        })
    except Exception as e:
        send_error(f"Failed to save analysis: {str(e)}")

def handle_load_history(_payload: Dict[str, Any]) -> None:
    """Load list of saved analyses"""
    try:
        from pathlib import Path
        import os
        
        home_dir = Path.home() / '.bioviz_local'
        if not home_dir.exists():
            send_response({"status": "ok", "history": []})
            return
            
        files = sorted(home_dir.glob("analysis_*.json"), key=os.path.getmtime, reverse=True)
        history = []
        
        for f in files:
            try:
                with open(f, 'r', encoding='utf-8') as json_file:
                    data = json.load(json_file)
                    history.append({
                        "filename": f.name,
                        "timestamp": data.get("timestamp"),
                        "pathway_id": data.get("pathway_id"),
                        "node_count": len(data.get("gene_expression", {})) if data.get("gene_expression") else 0
                    })
            except:
                continue
                
        send_response({"status": "ok", "history": history})
    except Exception as e:
        send_error(f"Failed to load history: {str(e)}")

def handle_load_analysis(payload: Dict[str, Any]) -> None:
    """Load a specific analysis file"""
    try:
        from pathlib import Path
        filename = payload.get("filename")
        if not filename:
            raise ValueError("Filename required")
            
        filepath = Path.home() / '.bioviz_local' / filename
        
        if not filepath.exists():
            raise FileNotFoundError(f"File {filename} not found")
            
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        send_response({
            "status": "ok",
            "data": data
        })
    except Exception as e:
        send_error(f"Failed to load analysis: {str(e)}")


# ========================
# AI Chat Handlers (Logic Lock)
# ========================

def handle_chat(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Process a user query through the AI Logic Lock system.
    
    Payload:
        query: str - The user's message
        history: list - Optional conversation history
        context: dict - Optional context (e.g., current gene_expression data)
    
    Returns:
        AIAction as dict with type: CHAT, EXECUTE, or PROPOSAL
    """
    try:
        from ai_core import process_query
        
        query = payload.get("query", "")
        history = payload.get("history", [])
        context = payload.get("context", {})
        
        if not query:
            return {"status": "error", "message": "Query is required"}
        
        print(f"[BioCore] Processing chat query: {query[:50]}...", file=sys.stderr)
        action = process_query(query, history, context)
        
        print(f"[BioCore] AI action type: {action.type}, content: {action.content[:50] if action.content else 'None'}...", file=sys.stderr)
        
        # Convert Pydantic model to dict
        action_dict = action.model_dump()
        print(f"[BioCore] Serialized action: {str(action_dict)[:100]}...", file=sys.stderr)
        
        return {
            "status": "ok",
            "cmd": "CHAT",  # Add cmd field for frontend
            **action_dict
        }
        
    except ImportError as e:
        error_msg = f"AI module not available: {str(e)}. Please install openai and pydantic."
        print(f"[BioCore] Import error: {error_msg}", file=sys.stderr)
        return {
            "status": "ok",
            "cmd": "CHAT",
            "type": "CHAT",
            "content": error_msg
        }
    except Exception as e:
        error_msg = f"AI error: {str(e)}"
        print(f"[BioCore] Error: {error_msg}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr) # Print traceback to stderr
        return {
            "status": "ok",
            "cmd": "CHAT",
            "type": "CHAT",
            "content": f"Sorry, I encountered an error: {str(e)}"
        }


def handle_chat_confirm(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Confirm and execute a previously proposed action.
    
    Payload:
        proposal_id: str - UUID of the proposal to confirm
        context: dict - Optional context
    """
    try:
        from ai_core import execute_proposal
        
        proposal_id = payload.get("proposal_id")
        context = payload.get("context", {})
        
        if not proposal_id:
            return {"status": "error", "message": "proposal_id is required"}
        
        action = execute_proposal(proposal_id, context)
        
        return {
            "status": "ok",
            **action.model_dump()
        }
        
    except Exception as e:
        return {
            "status": "error",
            "message": f"Error confirming proposal: {str(e)}"
        }

def _get_kegg_participants(pathway_id: str) -> List[str]:
    """Fetch gene symbols for a KEGG pathway."""
    import requests
    from typing import List
    try:
        # Step 1: Get Entrez IDs
        kid = pathway_id
        if ':' in kid: kid = kid.split(':')[-1]
        url = f"https://rest.kegg.jp/link/hsa/{kid}"
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        
        entrez_ids = []
        for line in response.text.strip().split('\n'):
            if '\t' in line:
                entrez_id = line.split('\t')[1]
                entrez_ids.append(entrez_id)
        
        if not entrez_ids:
            return []
            
        # Step 2: Get symbols for these Entrez IDs (batch of 100)
        symbols = []
        for i in range(0, len(entrez_ids), 100):
            batch = entrez_ids[i:i+100]
            list_url = f"https://rest.kegg.jp/list/{'+'.join(batch)}"
            list_response = requests.get(list_url, timeout=10)
            list_response.raise_for_status()
            
            for line in list_response.text.strip().split('\n'):
                if '\t' in line:
                    # hsa:123  SYMBOL, full name
                    symbol_part = line.split('\t')[1]
                    symbol = symbol_part.split(',')[0].strip()
                    symbols.append(symbol)
                    
        logging.info(f"Retrieved {len(symbols)} genes from KEGG for {pathway_id}")
        return symbols
    except Exception as e:
        logging.warning(f"Failed to get KEGG participants: {e}")
        return []

def _get_wikipathways_participants(pathway_id: str) -> List[str]:
    """Fetch gene symbols for a WikiPathways pathway."""
    import requests
    from typing import List
    try:
        wid = pathway_id
        if '_' in wid: wid = wid.split('_')[0]
        # Use getXrefList with code 'H' for HGNC symbols
        url = f"https://webservice.wikipathways.org/getXrefList?pwId={wid}&code=H&format=json"
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        
        data = response.json()
        symbols = data.get('xrefs', [])
        
        logging.info(f"Retrieved {len(symbols)} genes from WikiPathways for {pathway_id}")
        return symbols
    except Exception as e:
        logging.warning(f"Failed to get WikiPathways participants: {e}")
        return []

def handle_chat_reject(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Reject a previously proposed action.
    
    Payload:
        proposal_id: str - UUID of the proposal to reject
    """
    try:
        from ai_core import reject_proposal
        
        proposal_id = payload.get("proposal_id")
        
        if not proposal_id:
            return {"status": "error", "message": "proposal_id is required"}
        
        action = reject_proposal(proposal_id)
        
        return {
            "status": "ok",
            **action.model_dump()
        }
        
    except Exception as e:
        return {
            "status": "error",
            "message": f"Error rejecting proposal: {str(e)}"
        }


# ========================
# AI Structured Prompt Handlers
# ========================

def handle_summarize_enrichment(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Summarize enrichment results using structured prompts."""
    try:
        from ai_tools import summarize_enrichment

        enrichment_data = payload.get("enrichment_data") or payload.get("enriched_terms") or payload.get("data") or {}
        volcano_data = payload.get("volcano_data") or payload.get("volcanoData")
        context = payload.get("context") or {}
        metadata = {
            "pathway": payload.get("pathway") or context.get("pathway") or {},
            "statistics": payload.get("statistics") or context.get("statistics") or {},
        }

        return summarize_enrichment(enrichment_data, volcano_data=volcano_data, metadata=metadata)
    except Exception as e:
        return {"status": "error", "message": f"Failed to summarize enrichment: {str(e)}"}


def handle_summarize_de(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Summarize differential expression results."""
    try:
        from ai_tools import summarize_de_genes

        volcano_data = payload.get("volcano_data") or payload.get("volcanoData") or []
        thresholds = payload.get("thresholds") or {}
        return summarize_de_genes(volcano_data, thresholds)
    except Exception as e:
        return {"status": "error", "message": f"Failed to summarize differential expression: {str(e)}"}


def handle_parse_filter(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Parse natural language filters into structured logic."""
    try:
        from ai_tools import parse_filter_query

        query = payload.get("query") or payload.get("text") or ""
        available_fields = payload.get("available_fields") or payload.get("columns") or []
        return parse_filter_query(query, available_fields)
    except Exception as e:
        return {"status": "error", "message": f"Failed to parse filter query: {str(e)}"}


def handle_generate_hypothesis(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Generate Phase 3 hypotheses with explicit disclaimers."""
    try:
        from ai_tools import generate_hypothesis

        significant_genes = payload.get("significant_genes") or payload.get("genes")
        pathways = payload.get("pathways") or payload.get("enriched_terms")
        volcano_data = payload.get("volcano_data") or payload.get("volcanoData")
        return generate_hypothesis(significant_genes, pathways=pathways, volcano_data=volcano_data)
    except Exception as e:
        return {"status": "error", "message": f"Failed to generate hypothesis: {str(e)}"}


def handle_discover_patterns(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Run exploratory pattern discovery prompts."""
    try:
        from ai_tools import discover_patterns

        expression_matrix = (
            payload.get("expression_matrix")
            or payload.get("expressionMatrix")
            or payload.get("volcano_data")
            or payload.get("volcanoData")
        )
        return discover_patterns(expression_matrix)
    except Exception as e:
        return {"status": "error", "message": f"Failed to discover patterns: {str(e)}"}


def handle_describe_visualization(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Describe visualization trends without causal claims."""
    try:
        from ai_tools import describe_visualization

        table_data = (
            payload.get("table_data")
            or payload.get("enrichment_data")
            or payload.get("volcano_data")
            or payload.get("volcanoData")
        )
        return describe_visualization(table_data)
    except Exception as e:
        return {"status": "error", "message": f"Failed to describe visualization: {str(e)}"}

def process_command(command_obj: Dict[str, Any]) -> None:
    """Route command to appropriate handler."""
    global CURRENT_REQUEST_ID, CURRENT_CMD
    try:
        # command_obj is already a dict from main
        cmd = command_obj.get("cmd", "").upper() # Use 'cmd' as per existing protocol
        payload = command_obj.get("payload", {}) # Use 'payload' for consistency
        request_id = command_obj.get("request_id")
        CURRENT_REQUEST_ID = str(request_id) if request_id is not None else None
        CURRENT_CMD = cmd or None

        logging.info(f"[CMD] Processing command: {cmd} (request_id={request_id})")

        # Handlers that return a dict (old style)
        return_handlers = {
            "HEARTBEAT": handle_heartbeat,
            "LOAD": handle_load,
            "ANALYZE": handle_analyze,
            "LOAD_PATHWAY": handle_load_pathway,
            "COLOR_PATHWAY": handle_color_pathway,
            "SEARCH_PATHWAY": handle_search_pathways,
            "DOWNLOAD_PATHWAY": handle_download_pathway,
            "LIST_TEMPLATES": handle_list_templates,
            # AI Chat handlers (Logic Lock)
            "CHAT": handle_chat,
            "CHAT_CONFIRM": handle_chat_confirm,
            "CHAT_REJECT": handle_chat_reject,
            # Structured prompt handlers
            "SUMMARIZE_ENRICHMENT": handle_summarize_enrichment,
            "SUMMARIZE_DE": handle_summarize_de,
            "PARSE_FILTER": handle_parse_filter,
            "GENERATE_HYPOTHESIS": handle_generate_hypothesis,
            "DISCOVER_PATTERNS": handle_discover_patterns,
            "DESCRIBE_VISUALIZATION": handle_describe_visualization,
        }
        
        if GSEA_AVAILABLE:
            from gsea_module import (
                handle_run_enrichr,
                handle_run_gsea,
                handle_get_gene_sets,
                handle_load_gmt,
                handle_export_csv
            )
            return_handlers["ENRICHR"] = handle_run_enrichr
            return_handlers["GSEA"] = handle_run_gsea
            return_handlers["GET_GENE_SETS"] = handle_get_gene_sets
            return_handlers["LOAD_GMT"] = handle_load_gmt
            return_handlers["EXPORT_CSV"] = handle_export_csv
            logging.debug("GSEA handlers available")
        else:
            logging.warning("GSEA handlers NOT available")
        
        # V2.0: Add Image handlers if available
        if IMAGE_AVAILABLE:
            return_handlers["UPLOAD_IMAGE"] = handle_upload_image
            return_handlers["ANALYZE_IMAGE"] = handle_analyze_image
            return_handlers["LIST_IMAGES"] = handle_list_images
            logging.debug("Image handlers available")
        else:
            logging.warning("Image handlers NOT available")
        
        # V2.0: Add Multi-sample handlers if available
        if MULTI_SAMPLE_AVAILABLE:
            return_handlers["LOAD_MULTI_SAMPLE"] = handle_load_multi_sample
            return_handlers["GET_SAMPLE_GROUPS"] = handle_get_sample_groups
            logging.debug("Multi-sample handlers available")
        else:
            logging.warning("Multi-sample handlers NOT available")
        
        # V2.0: Add DE Analysis handler
        try:
            from de_analysis import handle_de_analysis
            return_handlers["DE_ANALYSIS"] = handle_de_analysis
            logging.debug("DE Analysis handler registered")
        except ImportError as e:
            logging.warning(f"DE Analysis handler NOT available: {e}")
        
        # V2.0: Add Enrichment Framework v2.0 handlers
        return_handlers["ENRICH_RUN"] = handle_enrich_run
        return_handlers["GENE_SET_LIST"] = handle_gene_set_list
        return_handlers["LOAD_CUSTOM_GMT"] = handle_load_custom_gmt
        return_handlers["BATCH_ENRICH_RUN"] = handle_batch_enrich_run
        return_handlers["EXPORT_ENRICHMENT"] = handle_export_enrichment
        logging.debug("Enrichment Framework v2.0 handlers registered")
        
        # V3.0: Reactome visualization handlers
        return_handlers["LOAD_REACTOME_PATHWAY"] = handle_load_reactome_pathway
        return_handlers["SEARCH_REACTOME"] = handle_search_reactome
        logging.debug("Reactome v3.0 handlers registered")
        
        # V4.0: Unified pathway framework handlers
        return_handlers["LOAD_UNIFIED_PATHWAY"] = handle_load_unified_pathway
        return_handlers["SEARCH_PATHWAYS"] = handle_search_pathways
        return_handlers["SEARCH_AND_LOAD_PATHWAY"] = handle_search_and_load_pathway
        logging.debug("Unified Pathway v4.0 handlers registered")

        # Handlers that send response directly (new style)
        direct_send_handlers = {
            "SAVE_ANALYSIS": handle_save_analysis,
            "LOAD_HISTORY": handle_load_history,
            "LOAD_ANALYSIS": handle_load_analysis,
            "SAVE_DATA": handle_save_data,
        }
        
        # Check both handler dicts
        if cmd in return_handlers:
            logging.info(f"[CMD] Calling handler for: {cmd}")
            result = return_handlers[cmd](payload)
            logging.info(f"[CMD] Handler completed: {cmd}, status={result.get('status', 'unknown')}")
            send_response(result)
        elif cmd in direct_send_handlers:
            logging.info(f"[CMD] Calling direct handler for: {cmd}")
            direct_send_handlers[cmd](payload)
            logging.info(f"[CMD] Direct handler completed: {cmd}")
        else:
            logging.error(f"[CMD] Unknown command: {cmd}")
            send_error(
                f"Unknown command: {cmd}",
                details={"available_commands": list(return_handlers.keys()) + list(direct_send_handlers.keys())}
            )
            
    except json.JSONDecodeError as e:
        logging.error(f"[CMD] Invalid JSON: {e}")
        send_error(f"Invalid JSON: {str(e)}")
    except Exception as e:
        logging.exception(f"[CMD] System error: {e}")
        send_error(f"System error: {str(e)}", details={"traceback": traceback.format_exc()})
    finally:
        # Always clear context after handling one command to avoid leaking into later responses.
        CURRENT_REQUEST_ID = None
        CURRENT_CMD = None



def run():
    """
    Main daemon loop.
    
    IMPORTANT: This is an infinite loop that only exits when:
    1. stdin is closed (parent process exits)
    2. Process is killed
    
    DO NOT add any exit conditions that could cause premature termination.
    """
    sys.stderr.reconfigure(encoding='utf-8')
    # sys.stdout is used for JSON output, ensure it's unbuffered or flushed often
    # But for JSON communication, we usually print one line per message.
    
    print("[BioEngine] Engine started. Waiting for commands...", file=sys.stderr)
    print(f"[BioEngine] Python version: {sys.version}", file=sys.stderr)

    # Send startup confirmation
    send_response({"status": "ready", "message": "BioViz Engine initialized"})
    
    while True:
        try:
            # Blocking read from stdin
            # This will return empty string when stdin is closed
            line = sys.stdin.readline()
            
            # Check for EOF (parent process closed stdin)
            if not line:
                break
            
            # Strip whitespace and skip empty lines
            line = line.strip()
            if not line:
                continue
            
            # Parse JSON command
            try:
                payload = json.loads(line)
            except json.JSONDecodeError as e:
                send_response({
                    "status": "error",
                    "message": f"Invalid JSON: {str(e)}",
                    "received": line[:100]  # First 100 chars for debugging
                })
                continue
            
            # Process command and send response
            process_command(payload)
            # Response is sent within process_command
            
        except KeyboardInterrupt:
            # Graceful shutdown on Ctrl+C
            break
        except Exception as e:
            # Catch-all for unexpected errors
            # NEVER let the daemon crash
            send_response({
                "status": "error",
                "message": f"Unexpected error: {str(e)}",
                "traceback": traceback.format_exc()
            })



# --- KEGG Search & Download ---
def list_local_templates() -> List[Dict[str, Any]]:
    """
    Enumerate available pathway templates from user and bundled locations.
    User templates take priority (deduplicate by ID).
    """
    candidate_dirs = [
        Path.home() / '.bioviz_local' / 'templates',
        Path(__file__).parent.parent / 'assets' / 'templates',
        Path(sys.executable).parent / 'assets' / 'templates',
        Path(sys.executable).parent.parent / 'Resources' / 'assets' / 'templates',
        Path(sys.executable).parent.parent / 'Resources' / '_up_' / 'assets' / 'templates',
        Path.cwd() / 'assets' / 'templates',
        Path.cwd().parent / 'assets' / 'templates',
    ]

    seen: set[str] = set()
    templates: List[Dict[str, Any]] = []

    for folder in candidate_dirs:
        if not folder.exists() or not folder.is_dir():
            continue
        for tpl_file in sorted(folder.glob("*.json")):
            try:
                with open(tpl_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                tpl_id = data.get("id") or tpl_file.stem
                if tpl_id in seen:
                    continue  # user folder already provided this ID
                name = data.get("name") or tpl_id
                desc = data.get("description") or name
                types = data.get("types") or ['gene', 'protein', 'cell']
                templates.append({
                    "id": tpl_id,
                    "name": name,
                    "description": desc,
                    "path": str(tpl_file),
                    "types": types,
                })
                seen.add(tpl_id)
            except Exception as e:
                print(f"[BioEngine] Skip template {tpl_file}: {e}", file=sys.stderr)

    return templates

def search_kegg_pathways(query: str) -> List[Dict[str, str]]:
    """
    Search KEGG pathways by query string.
    Uses KEGG REST API: http://rest.kegg.jp/find/pathway/{query}
    """
    import urllib.request
    import urllib.parse
    
    query = urllib.parse.quote(query)
    url = f"http://rest.kegg.jp/find/pathway/{query}"
    
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
    req = urllib.request.Request(url, headers=headers)
    
    results = []
    try:
        with urllib.request.urlopen(req, timeout=10) as response:
            data = response.read().decode('utf-8')
            for line in data.strip().split('\n'):
                if not line: continue
                parts = line.split('\t')
                if len(parts) >= 2:
                    kegg_id = parts[0].replace('path:', '')
                    
                    # Handle Reference Pathways (map) -> Convert to Human (hsa)
                    if kegg_id.startswith('map'):
                        kegg_id = kegg_id.replace('map', 'hsa')
                    elif not kegg_id.startswith('hsa'):
                        # Skip other organisms or ko/ec
                        continue
                        
                    desc = parts[1]
                    # Clean description "Name - Homo sapiens (human)" -> "Name"
                    if ' - Homo sapiens' in desc:
                        desc = desc.split(' - Homo sapiens')[0]
                        
                    results.append({
                        "id": kegg_id,
                        "name": desc,
                        "description": desc # simple fallback
                    })
        return results
        return results
    except urllib.error.URLError as e:
        print(f"[BioEngine] Network error: {e}", file=sys.stderr)
        # Return a special error result or just empty list with logging
        # Since this returns a list, we might need to handle the error at the caller level 
        # OR just return empty list and let the frontend handle "No results"
        # Ideally, we should raise or return an error dict, but the signature is List[Dict].
        # For now, let's log it. 
        # Better: Modify signature to allow returning error? 
        # Existing caller `handle_search_pathways` expects list.
        # Let's return a "fake" result item indicating error if possible, OR just handle in `handle_search_pathways`.
        raise e # Let caller handle it
    except Exception as e:
        print(f"[BioEngine] Search failed: {e}", file=sys.stderr)
        return []

def kgml_to_json(kgml_content: str, pathway_id: str) -> Dict[str, Any]:
    """
    Parse KGML XML content into BioViz JSON template format.
    """
    import xml.etree.ElementTree as ET
    
    root = ET.fromstring(kgml_content)
    
    # Metadata
    title = root.get('title', pathway_id)
    image_url = root.get('image', '')
    
    nodes = []
    edges = []
    
    # Map entry ID (integer) to graphical ID (e.g. gene symbol)
    entry_id_map = {}
    
    # 1. Parse Entries (Nodes)
    for entry in root.findall('entry'):
        entry_id = entry.get('id')
        entry_type = entry.get('type')
        entry_name_raw = entry.get('name') # hsa:1234 hsa:5678
        
        # We only visualize genes and compounds for now
        # "map" type entries are links to other pathways, we exclude or handle differently?
        # For simplicity, let's keep gene, compound, ortholog
        valid_types = ['gene', 'compound', 'ortholog', 'group']
        if entry_type not in valid_types:
            continue
            
        graphics = entry.find('graphics')
        if graphics is None:
            continue
            
        x_str = graphics.get('x')
        y_str = graphics.get('y')
        
        if not x_str or not y_str:
            continue
            
        try:
            x = int(x_str)
            y = int(y_str)
        except ValueError:
            continue
            
        # Get Display Name
        # Graphics name often has the common name "TP53, P53..."
        label = graphics.get('name')
        if label:
            label = label.split(',')[0].replace('...', '')
        else:
            label = entry_name_raw.split(' ')[0] if entry_name_raw else entry_id
            
        # Heuristic Category Mapping
        category = "Gene"
        fg_color = graphics.get('fgcolor') # KGML color hint?
        
        if entry_type == 'compound':
            category = "Compound"
        elif entry_type == 'group':
            category = "Complex"
        
        # KGML coords are center-based. 
        # Use internal_id as unique ID to avoid ECharts "duplicate name" error
        # (Same gene can appear multiple times in a map)
        
        node = {
            "id": entry_id, # Unique ID (e.g. "12")
            "name": label, # Display Name (e.g. "AKT1")
            "kegg_id": entry_name_raw, 
            "x": x,
            "y": y,
            "category": category,
            "internal_id": entry_id 
        }
        nodes.append(node)
        entry_id_map[entry_id] = entry_id # Map to unique ID for edges
        
        # Handle Groups (Components)
        for comp in entry.findall('component'):
            comp_id = comp.get('id')
            # Edge case handling...
            
    # 2. Parse Relations (Edges)
    for rel in root.findall('relation'):
        entry1 = rel.get('entry1')
        entry2 = rel.get('entry2')
        rel_type = rel.get('type')
        
        source = entry_id_map.get(entry1)
        target = entry_id_map.get(entry2)
        
        # If one of the endpoints wasn't a valid node (e.g. a "map" link), skip
        if not source or not target:
            continue
            
        # Map KGML relation to our types
        # GErel: expression
        # PPrel: protein-protein interaction
        # PCrel: protein-compound
        
        relation_str = "interaction"
        subtype_el = rel.find('subtype')
        if subtype_el is not None:
             subtype_name = subtype_el.get('name')
             # activation, inhibition, phosphorylation, ubiquitination...
             if subtype_name in ['activation', 'expression', 'indirect effect']:
                 relation_str = "activation"
             elif subtype_name in ['inhibition', 'repression', 'dephosphorylation']:
                 relation_str = "inhibition"
             elif subtype_name in ['phosphorylation']:
                 relation_str = "phosphorylation"
             elif subtype_name in ['ubiquitination']:
                 relation_str = "ubiquitination"
             elif subtype_name in ['binding/association', 'complex']:
                 relation_str = "binding"
        
        edge = {
            "source": source,
            "target": target,
            "relation": relation_str
        }
        edges.append(edge)

    # Construct final JSON
    return {
        "id": pathway_id,
        "name": title,
        "description": f"Imported from KEGG: {title}",
        "nodes": nodes,
        "edges": edges,
        "categories": {
            "Gene": "#3498db",
            "Compound": "#f1c40f", 
            "Complex": "#9b59b6",
            "Unknown": "#95a5a6"
        }
    }


def download_kegg_pathway(pathway_id: str) -> Dict[str, Any]:
    """
    Download KGML for pathway, parse to JSON, and save to assets.
    """
    import urllib.request
    
    # 1. Fetch KGML
    url = f"http://rest.kegg.jp/get/{pathway_id}/kgml"
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
    req = urllib.request.Request(url, headers=headers)

    try:
        print(f"[BioEngine] DownloadingKGML: {url}", file=sys.stderr)
        with urllib.request.urlopen(req, timeout=15) as response:
            kgml_content = response.read().decode('utf-8')
    except Exception as e:
         return {"status": "error", "message": f"Failed to download KGML: {str(e)}"}
         
    # 2. Parse
    try:
        template_json = kgml_to_json(kgml_content, pathway_id)
    except Exception as e:
        return {"status": "error", "message": f"Failed to parse KGML: {str(e)}"}
        
    # 3. Save to assets/templates
    # Determine save path based on execution environment
    try:
        # Use user-writable location first
        user_dir = Path.home() / '.bioviz_local' / 'templates'
        user_dir.mkdir(parents=True, exist_ok=True)

        # Dev path (if exists) for convenience
        dev_dir = Path(__file__).parent.parent / 'assets' / 'templates'
        target_dir = user_dir if user_dir.exists() else dev_dir
        if dev_dir.exists():
            # still prefer user dir to avoid writing into bundle
            target_dir = user_dir

        file_path = target_dir / f"{pathway_id}.json"
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(template_json, f, indent=4, ensure_ascii=False)
            
        return {
            "status": "ok", 
            "message": f"Saved to {file_path}", 
            "path": str(file_path),
            "template": template_json
        }
            
    except Exception as e:
        return {"status": "error", "message": f"Failed to save content: {str(e)}"}


def handle_search_pathways(payload: Dict[str, Any]) -> Dict[str, Any]:
    query = payload.get("query", "")
    if not query:
        return {"status": "error", "message": "No query provided"}
    try:
        results = search_kegg_pathways(query)
        return {"status": "ok", "results": results}
    except Exception as e:
         return {
             "status": "error", 
             "message": f"Network Error: Failed to connect to KEGG. Please check your internet connection or use a VPN (KEGG is often blocked in China). Details: {str(e)}"
         }

def handle_download_pathway(payload: Dict[str, Any]) -> Dict[str, Any]:
    pid = payload.get("id", "")
    if not pid:
         return {"status": "error", "message": "No pathway ID provided"}
    return download_kegg_pathway(pid)


def handle_list_templates(_payload: Dict[str, Any]) -> Dict[str, Any]:
    """List local pathway templates from user folder and bundled assets."""
    return {"status": "ok", "templates": list_local_templates()}


# ============================================================================
# Enrichment Framework v2.0 Handlers
# ============================================================================

def handle_enrich_run(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Run enrichment analysis (ORA or GSEA) using new enrichment framework."""
    try:
        from enrichment.pipeline import EnrichmentPipeline
        
        method = payload.get('method', 'ORA').upper()
        genes = payload.get('genes', [])
        gene_set_source = payload.get('gene_set_source', 'reactome')
        species = payload.get('species', 'auto')
        custom_gmt_path = payload.get('custom_gmt_path')
        params = payload.get('parameters', {})
        
        if not genes:
            return {"status": "error", "message": "No genes provided"}
            
        pipeline = EnrichmentPipeline()
        
        if method == 'ORA':
            if isinstance(genes, dict):
                genes = list(genes.keys())
            
            result = pipeline.run_ora(
                gene_list=genes,
                gene_set_source=gene_set_source,
                species=species,
                custom_gmt_path=custom_gmt_path,
                p_cutoff=params.get('p_cutoff', 0.05),
                min_overlap=params.get('min_overlap', 3),
                fdr_method=params.get('fdr_method', 'fdr_bh')
            )
        
        elif method == 'GSEA':
            if isinstance(genes, list):
                return {"status": "error", "message": "GSEA requires ranked gene list"}
            
            result = pipeline.run_gsea(
                gene_ranking=genes,
                gene_set_source=gene_set_source,
                species=species,
                custom_gmt_path=custom_gmt_path,
                min_size=params.get('min_size', 5),
                max_size=params.get('max_size', 500),
                permutation_num=params.get('permutation_num', 1000)
            )
        
        # Advanced Intelligence Logic (Rule-based)
        all_res = []
        if method == 'ORA':
            all_res = result.get('results', [])
        else:
            all_res = result.get('up_regulated', []) + result.get('down_regulated', [])
            
        sig_pathways = [p for p in all_res if p.get('fdr', 1.0) < 0.05]
        
        # 1. Basic Stats
        sig_count = len(sig_pathways)
        
        # 2. Key Drivers (Genes in multiple pathways)
        gene_to_pathways = {}
        for p in sig_pathways:
            path_name = p.get('pathway_name', 'Unknown')
            # Handle hit_genes string or list
            hits = p.get('hit_genes', [])
            if isinstance(hits, str):
                hits = [g.strip() for g in hits.split(',') if g.strip()]
            for g in hits:
                if g not in gene_to_pathways: gene_to_pathways[g] = []
                gene_to_pathways[g].append(path_name)
        
        drivers = sorted([{"gene": g, "count": len(paths), "paths": paths[:3]} 
                        for g, paths in gene_to_pathways.items() if len(paths) >= 3], 
                        key=lambda x: x['count'], reverse=True)[:5]
        
        # 3. Orphan Significant Genes
        # We need the original genes to find orphans
        all_input_genes = []
        if isinstance(genes, dict):
            # genes is {symbol: logfc}
            all_input_genes = [{"gene": g, "logfc": val} for g, val in genes.items()]
        else:
            all_input_genes = [{"gene": g, "logfc": 0} for g in genes]
            
        # Filter for significant input genes (heuristic: top 50 by magnitude or p-value if we had it)
        # Since handle_enrich_run only gets 'genes', we use magnitude of logfc if provided
        sig_input = sorted([g for g in all_input_genes if abs(g['logfc']) > 1.0], 
                         key=lambda x: abs(x['logfc']), reverse=True)[:30]
        
        # A gene is an orphan if it's significant in DE but doesn't appear in ANY of the top 10 significant pathways
        top_10_pathway_genes = set()
        for p in sig_pathways[:10]:
            hits = p.get('hit_genes', [])
            if isinstance(hits, str):
                hits = [g.strip() for g in hits.split(',') if g.strip()]
            top_10_pathway_genes.update(hits)
            
        orphans = [g['gene'] for g in sig_input if g['gene'] not in top_10_pathway_genes][:5]
        
        # 4. Antagonistic Pathways (Identify if any pathways show balanced mixed regulation)
        antagonistic_paths = []
        for p in sig_pathways:
            hits = p.get('hit_genes', [])
            if isinstance(hits, str): hits = [g.strip() for g in hits.split(',') if g.strip()]
            
            up_c, down_c = 0, 0
            for g in hits:
                lfc = genes.get(g, 0) if isinstance(genes, dict) else 0
                if lfc > 0.5: up_c += 1
                elif lfc < -0.5: down_c += 1
            
            is_antag = (up_c > 0 and down_c > 0 and (min(up_c, down_c) / max(1, len(hits)) > 0.25))
            if is_antag:
                p['is_antagonistic'] = True
                antagonistic_paths.append(p.get('pathway_name', 'Unknown'))
            else:
                p['is_antagonistic'] = False

        # 5. Redundancy / Hierarchy (Group pathways with similar names)
        keyword_counts = {}
        stop_words = {'signaling', 'pathway', 'by', 'of', 'and', 'the', 'in', 'r-hsa', 'hsa'}
        for p in sig_pathways[:15]:
            words = set(p.get('pathway_name', '').lower().replace('-', ' ').replace('_', ' ').split())
            cleaned = words - stop_words
            for w in cleaned:
                if len(w) > 3:
                    keyword_counts[w] = keyword_counts.get(w, 0) + 1
        
        redundant_themes = [k for k, v in keyword_counts.items() if v >= 4][:3]

        # 6. Silent Pathway Members (Significant enrichment but few total hits)
        silent_paths = []
        for p in sig_pathways[:5]:
            # If overlap is very low (e.g. < 10% of pathway) but still p < 0.05
            # We need the denominator from overlap_ratio like "10/700"
            overlap_str = p.get('overlap_ratio', '0/1')
            try:
                hits_c, total_c = map(int, overlap_str.split('/'))
                if hits_c / total_c < 0.05 and hits_c > 0:
                    silent_paths.append(p.get('pathway_name', 'Unknown'))
            except: pass

        # 7. Leading-Edge Trends (For GSEA - consistency check)
        # (Already handled by stat logic, but we can flag it)

        # Build Insights
        insights = []
        if sig_count > 0:
            insights.append(f"Found {sig_count} pathways with high statistical confidence.")
            
            if redundant_themes:
                themes_str = ", ".join([t.capitalize() for t in redundant_themes])
                insights.append(f"Systemic Redundancy: Multiple significant hits relate to {themes_str} processes.")
            
            if drivers:
                insights.append(f"Key Drivers: {', '.join([d['gene'] for d in drivers])} are influencing multiple systems.")
            
            if antagonistic_paths:
                insights.append(f"Antagonistic Regulation: Pathways like '{antagonistic_paths[0]}' show balanced mixed regulation.")
            
            if orphans:
                insights.append(f"Orphan Genes: {', '.join(orphans)} show high impact but aren't mapped to top pathways.")
            
            if silent_paths:
                insights.append(f"Precise Regulation: '{silent_paths[0]}' is significant despite only a tiny fraction of its members being active.")
        else:
            # Baseline / Integrity Layer
            input_size = len(genes) if isinstance(genes, list) else len(genes.keys())
            if input_size < 30:
                insights.append("Data Sparsity: Input gene list is too short for reliable biological enrichment.")
            else:
                insights.append("Discrete Signals: High-impact genes detected but they do not cluster into known biological pathways.")

        result['intelligence_report'] = {
            "summary": insights[0] if insights else "Analysis complete.",
            "full_details": insights,
            "sig_count": sig_count,
            "drivers": drivers,
            "orphans": orphans,
            "antagonistic": antagonistic_paths[:3],
            "redundant_themes": redundant_themes,
            "silent_paths": silent_paths[:3]
        }
        
        result['standard_summary'] = insights[0] if insights else "Analysis complete."
        return result
        
    except Exception as e:
        logging.error(f"Enrichment analysis failed: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}


def handle_gene_set_list(payload: Dict[str, Any]) -> Dict[str, Any]:
    """List available gene set sources."""
    try:
        from enrichment.sources import GeneSetSourceManager
        
        species = payload.get('species', 'human')
        manager = GeneSetSourceManager()
        
        sources = manager.get_available_sources(species)
        
        return {
            "status": "ok",
            "sources": sources,
            "species": species
        }
        
    except Exception as e:
        logging.error(f"Failed to list gene sets: {e}")
        return {"status": "error", "message": str(e)}


def handle_load_custom_gmt(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Load and validate a custom GMT file."""
    try:
        from gene_set_utils import load_gmt, get_gene_set_stats, validate_gmt
        
        gmt_path = payload.get('path')
        gmt_content = payload.get('content')  # For direct text upload
        
        if not gmt_path and not gmt_content:
            return {"status": "error", "message": "No GMT file path or content provided"}
        
        # If content is provided directly (from frontend drag-drop)
        if gmt_content:
            import tempfile
            import os
            
            # Save to temp file
            temp_dir = Path.home() / '.bioviz' / 'cache' / 'custom_gmt'
            temp_dir.mkdir(parents=True, exist_ok=True)
            temp_file = temp_dir / f"custom_{hash(gmt_content[:100])}.gmt"
            
            with open(temp_file, 'w', encoding='utf-8') as f:
                f.write(gmt_content)
            
            gmt_path = str(temp_file)
        
        # Validate and load
        validation = validate_gmt(gmt_path)
        if not validation['valid']:
            return {"status": "error", "message": validation['error']}
        
        # Load gene sets
        gene_sets = load_gmt(gmt_path)
        stats = get_gene_set_stats(gene_sets)
        
        # Cache the loaded GMT for later use
        from enrichment.sources import GeneSetSourceManager
        manager = GeneSetSourceManager()
        manager.register_custom_gmt(gmt_path, gene_sets)
        
        return {
            "status": "ok",
            "path": gmt_path,
            "stats": {
                "geneSets": stats['num_sets'],
                "totalGenes": stats['total_genes'],
                "avgSetSize": stats['avg_set_size'],
                "fileName": Path(gmt_path).name
            }
        }
        
    except Exception as e:
        logging.error(f"Failed to load custom GMT: {e}")
        return {"status": "error", "message": str(e)}


def handle_batch_enrich_run(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Run batch enrichment analysis on multiple gene lists."""
    try:
        from enrichment.batch import run_batch_enrichment
        
        gene_lists = payload.get('gene_lists', {})
        gene_set_source = payload.get('gene_set_source', 'reactome')
        species = payload.get('species', 'human')
        method = payload.get('method', 'ORA')
        parameters = payload.get('parameters', {})
        
        if not gene_lists:
            return {"status": "error", "message": "No gene lists provided"}
        
        logging.info(f"Running batch enrichment: {len(gene_lists)} samples, method={method}")
        
        result = run_batch_enrichment(
            gene_lists=gene_lists,
            gene_set_source=gene_set_source,
            species=species,
            method=method,
            parameters=parameters
        )
        
        return result
        
    except Exception as e:
        logging.error(f"Batch enrichment failed: {e}")
        return {"status": "error", "message": str(e)}


def handle_export_enrichment(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Export enrichment results to file."""
    try:
        from enrichment.batch import export_batch_results
        import tempfile
        from pathlib import Path
        
        results = payload.get('results', {})
        format = payload.get('format', 'csv')
        output_path = payload.get('output_path')
        
        if not results:
            return {"status": "error", "message": "No results to export"}
        
        # If no output path, create temp file
        if not output_path:
            temp_dir = Path(tempfile.gettempdir())
            ext = {'xlsx': 'xlsx', 'csv': 'csv', 'json': 'json'}.get(format, 'csv')
            output_path = str(temp_dir / f"enrichment_results.{ext}")
        
        saved_path = export_batch_results(
            batch_results=results,
            output_path=output_path,
            format=format
        )
        
        return {
            "status": "ok",
            "path": saved_path,
            "format": format
        }
        
    except Exception as e:
        logging.error(f"Export failed: {e}")
        return {"status": "error", "message": str(e)}


def handle_load_reactome_pathway(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Load a Reactome pathway for visualization."""
    try:
        from reactome.client import ReactomeClient, convert_reactome_to_template
        
        pathway_id = payload.get('pathway_id')
        if not pathway_id:
            return {"status": "error", "message": "No pathway_id provided"}
        
        logging.info(f"Loading Reactome pathway: {pathway_id}")
        
        client = ReactomeClient()
        
        # Get pathway info
        pathway_info = client.get_pathway_info(pathway_id)
        if not pathway_info:
            return {"status": "error", "message": f"Pathway not found: {pathway_id}"}
        
        # Get diagram data
        diagram_data, entity_map = client.get_pathway_diagram(pathway_id)
        
        # Get participating genes for expression overlay
        gene_list = client.get_pathway_participants(pathway_id)
        
        # Convert to BioViz template format
        template = convert_reactome_to_template(diagram_data, entity_map, pathway_info)
        template['genes'] = gene_list
        
        return {
            "status": "ok",
            "pathway": template,
            "pathway_id": pathway_id,
            "name": pathway_info.get('displayName', ''),
            "source": "reactome",
            "gene_count": len(gene_list)
        }
        
    except Exception as e:
        logging.error(f"Failed to load Reactome pathway: {e}")
        return {"status": "error", "message": str(e)}


def handle_search_reactome(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Search for Reactome pathways."""
    try:
        from reactome.client import ReactomeClient
        
        query = payload.get('query', '')
        species = payload.get('species', 'Homo sapiens')
        limit = payload.get('limit', 20)
        
        if not query:
            return {"status": "error", "message": "No search query provided"}
        
        client = ReactomeClient()
        results = client.search_pathways(query, species, limit)
        
        return {
            "status": "ok",
            "results": results,
            "count": len(results)
        }
        
    except Exception as e:
        logging.error(f"Reactome search failed: {e}")
        return {"status": "error", "message": str(e)}


def handle_load_unified_pathway(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Load a pathway using the unified adapter registry."""
    try:
        from pathway import AdapterRegistry
        
        source = payload.get('source', 'kegg')
        pathway_id = payload.get('pathway_id')
        
        if not pathway_id:
            return {"status": "error", "message": "No pathway_id provided"}
        
        logging.info(f"Loading unified pathway: {source}/{pathway_id}")
        
        pathway = AdapterRegistry.load_pathway(source, pathway_id)
        
        if not pathway:
            return {"status": "error", "message": f"Pathway not found: {source}/{pathway_id}"}
        
        return {
            "status": "ok",
            "pathway": pathway.to_dict(),
            "source": source,
            "pathway_id": pathway_id,
            "name": pathway.name,
            "gene_count": len(pathway.genes)
        }
        
    except Exception as e:
        logging.error(f"Failed to load unified pathway: {e}")
        return {"status": "error", "message": str(e)}


def handle_search_pathways(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Search pathways across all registered sources."""
    try:
        from pathway import AdapterRegistry
        
        query = payload.get('query', '')
        species = payload.get('species', 'human')
        sources = payload.get('sources')  # Optional list of sources to search
        limit = payload.get('limit', 10)
        
        if not query:
            return {"status": "error", "message": "No search query provided"}
        
        if sources:
            # Search specific sources
            results = {}
            for source in sources:
                adapter = AdapterRegistry.get(source)
                if adapter:
                    results[source] = [
                        {'id': p.id, 'name': p.name, 'source': p.source, 'species': p.species}
                        for p in adapter.search(query, species, limit)
                    ]
        else:
            # Search all sources
            all_results = AdapterRegistry.search_all(query, species, limit)
            results = {
                src: [{'id': p.id, 'name': p.name, 'source': p.source, 'species': p.species} for p in pathways]
                for src, pathways in all_results.items()
            }
        
        total_count = sum(len(v) for v in results.values())
        
        return {
            "status": "ok",
            "results": results,
            "total_count": total_count,
            "sources_searched": list(results.keys())
        }
        
    except Exception as e:
        logging.error(f"Pathway search failed: {e}")
        return {"status": "error", "message": str(e)}


def handle_search_and_load_pathway(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Search and load pathway from any source with template caching.
    
    Strategy:
    1. Extract pathway ID if present in name
    2. Check template manager (bundled + cache)
    3. If not found, download from API and cache
    4. If download fails, suggest alternative sources
    """
    try:
        pathway_name = payload.get('pathway_name', '')
        source = payload.get('source', 'reactome').lower()
        species = payload.get('species', 'human')
        
        if not pathway_name:
            return {"status": "error", "message": "No pathway name provided"}
        
        logging.info(f"Search and load pathway: '{pathway_name}' from {source}")
        
        # Step 1: Extract pathway ID if embedded in name
        pathway_id = None
        import re
        
        if source == 'reactome':
            # Format: "Pathway Name R-HSA-1234567"
            match = re.search(r'(R-HSA-\d+)', pathway_name)
            if match:
                pathway_id = match.group(1)
                logging.info(f"Extracted Reactome ID: {pathway_id}")
        elif source == 'kegg':
            # Format: "hsa00010" or pathway name
            match = re.search(r'(hsa\d{5})', pathway_name)
            if match:
                pathway_id = match.group(1)
        elif source == 'wikipathways':
            # Format: "WP530"
            match = re.search(r'(WP\d+)', pathway_name)
            if match:
                pathway_id = match.group(1)
        
        # Step 2: Try to get from template manager (bundled or cached)
        if pathway_id and TEMPLATE_MANAGER:
            template, source_type = TEMPLATE_MANAGER.get_template(pathway_id, source)
            if template:
                logging.info(f"Template loaded from {source_type}: {pathway_id}")
                
                # Check if template has actual content (nodes or genes)
                has_nodes = template.get('nodes') and len(template.get('nodes', [])) > 0
                has_genes = template.get('genes') and len(template.get('genes', [])) > 0
                
                if has_nodes:
                    # Template has nodes (visual diagram), use it
                    return {
                        "status": "ok",
                        "pathway": template,
                        "pathway_id": pathway_id,
                        "pathway_name": template.get('name') or template.get('title', pathway_name),
                        "source": source,
                        "from_cache": source_type,
                        "gene_count": len(template.get('genes', []))
                    }
                else:
                    # Template exists but is empty - will try auto-layout later
                    logging.warning(f"Template {pathway_id} is empty (no nodes/genes), will try auto-layout")
                    template = None  # Reset to trigger auto-layout
        
        
        # Step 3: Not in cache - need to download from API
        logging.info(f"Template not cached, attempting download from {source} API")
        
        template = None
        downloaded_pathway_id = pathway_id
        
        if source == 'reactome':
            template = _download_reactome_pathway(pathway_name, pathway_id, species)
            if template:
                downloaded_pathway_id = template.get('id', pathway_id)
        elif source == 'kegg':
            # KEGG templates should already be bundled
            logging.warning(f"KEGG pathway {pathway_id} not found in bundled templates")
        elif source == 'wikipathways':
            logging.info("WikiPathways API download not yet implemented")
        elif source == 'go_bp':
            # GO BP pathways don't have diagrams
            return {
                "status": "info",
                "message": "GO Biological Process pathways do not have visual diagrams",
                "suggest_external": f"https://www.ebi.ac.uk/QuickGO/term/{pathway_id or pathway_name}"
            }
        
        # Step 4: If download succeeded, check for content
        if template and downloaded_pathway_id and TEMPLATE_MANAGER:
            # Check if downloaded template has actual content (nodes or genes)
            has_nodes = template.get('nodes') and len(template.get('nodes', [])) > 0
            has_genes = template.get('genes') and len(template.get('genes', [])) > 0
            
            if has_nodes:
                # Template has nodes (visual diagram), cache and return
                TEMPLATE_MANAGER.save_to_cache(downloaded_pathway_id, source, template)
                return {
                    "status": "ok",
                    "pathway": template,
                    "pathway_id": downloaded_pathway_id,
                    "pathway_name": template.get('name') or template.get('title', pathway_name),
                    "source": source,
                    "from_cache": "downloaded",
                    "gene_count": len(template.get('genes', []))
                }
            else:
                logging.warning(f"Downloaded template {downloaded_pathway_id} is empty, falling through to auto-layout")
                template = None  # Reset to trigger auto-layout
        
        # Step 5: Try to get gene list for auto-layout if no template found
        resolved_id = downloaded_pathway_id or pathway_id
        gene_list = None
        note = "Diagram automatically generated from gene list using STRING PPI interactions."
        
        if not template and resolved_id:
            if source == 'reactome':
                try:
                    from reactome.client import ReactomeClient
                    client = ReactomeClient()
                    gene_list = client.get_pathway_participants(resolved_id)
                    
                    # Also try to get sub-pathways for the "Downstream Suggestion"
                    pathway_info = client.get_pathway_info(resolved_id)
                    if pathway_info and pathway_info.get('hasEvent'):
                        sub_events = pathway_info['hasEvent']
                        sub_pathways = [e['displayName'] for e in sub_events if e.get('className') == 'Pathway'][:3]
                        if sub_pathways:
                            note += f" Potential downstream sub-pathways: {', '.join(sub_pathways)}."
                except Exception as e:
                    logging.warning(f"Failed to get gene list/info from Reactome: {e}")
            elif source == 'kegg':
                gene_list = _get_kegg_participants(resolved_id)
            elif source == 'wikipathways':
                gene_list = _get_wikipathways_participants(resolved_id)
                
        # Step 6: Auto-generate diagram if we have gene list
        if gene_list and len(gene_list) > 0:
            try:
                logging.info(f"Generating auto-layout for {len(gene_list)} genes...")
                from pathway.auto_layout import PathwayAutoLayoutEngine
                
                engine = PathwayAutoLayoutEngine(layout_algorithm='force')
                template = engine.generate_diagram(
                    genes=gene_list,
                    pathway_name=pathway_name,
                    pathway_id=resolved_id,
                    source=source,
                    species=species.title()
                )
                
                if template:
                    # Update note with downstream suggestions
                    template['metadata']['note'] = note
                    
                    # Cache the auto-generated template
                    if resolved_id and TEMPLATE_MANAGER:
                        TEMPLATE_MANAGER.save_to_cache(resolved_id, source, template)
                    
                    return {
                        "status": "ok",
                        "pathway": template,
                        "pathway_id": resolved_id,
                        "pathway_name": pathway_name,
                        "source": source,
                        "from_cache": "auto_generated",
                        "gene_count": len(gene_list),
                        "source_url": template.get('metadata', {}).get('source_url')
                    }
            except Exception as e:
                logging.error(f"Auto-layout generation failed: {e}")

        # Step 7: Everything failed - suggest alternative sources
        if TEMPLATE_MANAGER:
            alternatives = TEMPLATE_MANAGER.search_across_sources(pathway_name)
            if alternatives:
                suggestions = []
                for alt_source, pathway_ids in alternatives.items():
                    suggestions.append(f"{alt_source}: {len(pathway_ids)} matches")
                
                return {
                    "status": "not_found",
                    "message": f"Pathway not available in {source}",
                    "alternatives": alternatives,
                    "suggestion_message": f"Found in other sources: {', '.join(suggestions)}"
                }
        
        # No alternatives found
        return {
            "status": "error",
            "message": f"Pathway '{pathway_name}' not found in {source} and no alternatives available"
        }
    
    except Exception as e:
        logging.error(f"Search and load pathway failed: {e}")
        traceback.print_exc()
        return {"status": "error", "message": str(e)}


def _download_reactome_pathway(pathway_name: str, pathway_id: Optional[str], species: str) -> Optional[Dict]:
    """
    Download pathway from Reactome API.
    
    Returns:
        Template dict or None if failed
    """
    try:
        from reactome.client import ReactomeClient, convert_reactome_to_template
        
        client = ReactomeClient()
        
        # If we have ID, use it directly
        if pathway_id:
            pathway_info = client.get_pathway_info(pathway_id)
            
            # API may return a list
            if isinstance(pathway_info, list):
                pathway_info = pathway_info[0] if pathway_info else {}
            
            if not pathway_info:
                return None
            
            # Get diagram and genes
            diagram_data, entity_map = client.get_pathway_diagram(pathway_id)
            gene_list = client.get_pathway_participants(pathway_id)
            
            # Build template
            pathway_name_display = pathway_info.get('displayName', '') or pathway_info.get('name', pathway_name)
            species_info = pathway_info.get('species', {})
            if isinstance(species_info, list):
                species_info = species_info[0] if species_info else {}
            species_name = species_info.get('displayName', 'Human') if isinstance(species_info, dict) else 'Human'
            
            template = {
                'id': pathway_id,
                'name': pathway_name_display,
                'source': 'reactome',
                'species': species_name,
                'nodes': [],
                'edges': [],
                'width': 1000,
                'height': 800,
                'genes': gene_list if isinstance(gene_list, list) else []
            }
            
            # Add diagram if available
            if diagram_data and isinstance(diagram_data, dict) and diagram_data.get('nodes'):
                template = convert_reactome_to_template(diagram_data, entity_map, pathway_info)
                template['genes'] = gene_list if isinstance(gene_list, list) else []
            
            return template
        
        # No ID - search by name
        species_name = 'Homo sapiens' if species == 'human' else species.title()
        search_results = client.search_pathways(pathway_name, species_name, limit=5)
        
        if not search_results:
            return None
        
        # Find best match
        best_match = None
        pathway_name_lower = pathway_name.lower()
        for result in search_results:
            if pathway_name_lower in result.get('name', '').lower():
                best_match = result
                break
        
        if not best_match:
            best_match = search_results[0]
        
        # Recursively download using the found ID
        found_id = best_match.get('stId') or best_match.get('id')
        return _download_reactome_pathway(pathway_name, found_id, species)
    
    except Exception as e:
        logging.error(f"Failed to download Reactome pathway: {e}")
        return None


if __name__ == "__main__":
    run()





